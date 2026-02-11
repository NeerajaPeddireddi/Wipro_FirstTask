import pandas as pd
from openpyxl import load_workbook
df=pd.read_excel("sales_data.xlsx",sheet_name="Sheet1")
df["Total"]=df["Quantity"]*df["Price"]
df.to_excel("sales_data.xlsx",index=False)
print("File Saved Successfully")

#loading existing worbook
wb = load_workbook("sales_data.xlsx")
#select sheet
ws = wb["Sheet1"]
#add header for total column
ws['D1']="Total"
# Loop through rows and calculate total
for row in range(2,ws.max_row+1):
    quantity=ws[f"B{row}"].value
    price=ws[f"C{row}"].value
    ws[f"D{row}"]=quantity*price
wb.save("sales_data.xlsx")
print("File Saved Successfully")