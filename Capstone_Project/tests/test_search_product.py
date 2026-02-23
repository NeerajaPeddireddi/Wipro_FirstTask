import csv
import os

import pytest
from selenium.webdriver.common.by import By

from Capstone_Project.conftest import logger
from Capstone_Project.pages.Search_Product import SearchProduct

from openpyxl import load_workbook
def read_search_data_excel(file_name):
    base_path = os.path.join(os.path.dirname(__file__), "../data")
    file_path = os.path.join(base_path, file_name)
    logger.info(f"Reading search data from Excel file: {file_path}")

    workbook = load_workbook(file_path)
    sheet = workbook["searchproduct_data"]   # Sheet name

    data = []
    headers = []

    # Get headers (first row)
    for cell in sheet[1]:
        headers.append(cell.value)

    # Get rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)
    logger.info(f"Total test data rows loaded: {len(data)}")
    return data
search_test_data = read_search_data_excel("data.xlsx")
@pytest.mark.order(3)
@pytest.mark.parametrize("data", search_test_data)
def test_search_product(setup,data):
    logger.info("========== STARTING SEARCH PRODUCT TEST ==========")
    driver = setup
    search_page=SearchProduct(driver)


    # Go to Shop
    logger.info("Navigating to Shop page")
    search_page.go_to_shop()
    current_url = driver.current_url
    logger.info(f"Current URL after navigation: {current_url}")
    assert "shop" in driver.current_url.lower(), (
        f"Shop page did not open.\n"
        f"Current URL: {driver.current_url}"
    )
    #taing from csv
    product_name = data["product_name"]
    logger.info(f"Searching for product: {product_name}")

    # Search product
    search_page.search_product(product_name)
    logger.info("Search submitted successfully")
    driver.implicitly_wait(3)
    # Validate search box contains entered value
    search_box_value = driver.find_element(
        *search_page.search_box_value
    ).get_attribute("value")

    assert search_box_value == product_name, (
        f"Search box value mismatch!\n"
        f"Expected: {product_name}\n"
        f"Actual: {search_box_value}"
    )

    try:
        logger.info("Validating search results presence")

        result = search_page.validate_search_results()

        assert isinstance(result, bool), (
            "validate_search_results() should return True or False"
        )

        # -------- POSITIVE CASE --------
        if result:
            logger.info("Search results found")

            # Assert result count > 0 (if method available)
            results_count = search_page.get_result_count()
            assert results_count > 0, (
                "Search results validation returned True, "
                "but result count is 0"
            )

            old_url = driver.current_url

            search_page.click_search_result()

            # Assert URL changed after clicking product
            assert driver.current_url != old_url, (
                "URL did not change after clicking search result"
            )

            product_title = driver.find_element(
                *search_page.product_title
            ).text

            assert product_title != "", "Product title is empty"

            assert product_name.lower() in product_title.lower(), (
                f"Wrong product page opened!\n"
                f"Expected: {product_name}\n"
                f"Actual Title: {product_title}\n"
                f"URL: {driver.current_url}"
            )

        # -------- NEGATIVE CASE --------
        else:
            logger.info("No products found - Validating negative scenario")

            no_result_message = search_page.get_no_result_message()

            assert no_result_message is not None and no_result_message != "", (
                "No result message is empty or not displayed"
            )

            assert "nothing found" in no_result_message.lower(), (
                f"Expected 'Sorry, nothing found.' message not displayed.\n"
                f"Actual message: {no_result_message}"
            )

    except Exception as e:
        pytest.fail(f"Unexpected error during search validation: {str(e)}")

    logger.info("========== SEARCH PRODUCT TEST COMPLETED ==========")