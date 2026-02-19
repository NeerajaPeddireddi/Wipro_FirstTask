import pytest
from Capstone_Project.conftest import logger
from Capstone_Project.pages.Order_product import OrderProduct
import os

from openpyxl import load_workbook
def read_order_data_excel(file_name):
    base_path = os.path.join(os.path.dirname(__file__), "../data")
    file_path = os.path.join(base_path, file_name)
    logger.info(f"Reading place order data from Excel file: {file_path}")
    workbook = load_workbook(file_path)
    sheet = workbook["placeorder_data"]   # Sheet name

    data = []
    headers = []

    # Get headers (first row)
    for cell in sheet[1]:
        headers.append(cell.value)

    # Get rows
    for row in sheet.iter_rows(min_row=2, values_only=True):

        row_data = dict(zip(headers, row))

        numeric_fields = ["phone", "zipcode"]

        for field in numeric_fields:
            if isinstance(row_data.get(field), float):
                row_data[field] = str(int(row_data[field]))

        data.append(row_data)
    logger.info(f"Total test data rows loaded: {len(data)}")
    return data


order_test_data = read_order_data_excel("data.xlsx")
@pytest.mark.order(6)
@pytest.mark.parametrize("data", order_test_data)
def test_place_order_cod(setup, data):
    logger.info("========== PLACE ORDER LOGIN TEST ==========")
    driver = setup
    order = OrderProduct(driver)

    # Step 1: Go to Shop
    logger.info("Navigating to Shop page")
    order.go_to_shop()
    current_url = driver.current_url
    logger.info(f"Current URL after navigation: {current_url}")

    # Step 2: Add product to basket
    logger.info("Adding product to basket")
    order.add_product_to_basket()

    # Step 3: View basket
    logger.info("Viewing basket")
    order.click_view_basket()

    # Step 4: Proceed to checkout
    logger.info("Proceeding to checkout")
    order.click_proceed_to_checkout()

    # Step 5: Fill billing details
    logger.info("Filling billing details")
    order.enter_first_name(data["first_name"])
    order.enter_last_name(data["last_name"])
    order.enter_company_name(data["company"])
    order.enter_email(data["email"])
    order.enter_phone(data["phone"])

    # Step 6: Select country
    logger.info(f"Selecting country: {data['country']}")
    order.select_country_using_search(data["country"])

    # Step 7: Address details
    logger.info("Entering address details")
    order.enter_address(data["address"])
    order.enter_city(data["city"])
    order.select_state(data["state"])
    order.enter_zipcode(data["zipcode"])

    # Step 8: Order notes
    logger.info("Entering order notes")
    order.enter_order_notes(data["order_notes"])

    # Step 9: Select payment method
    logger.info(f"Selecting payment method: {data['payment_method']}")
    order.select_payment_method(data["payment_method"])

    # Step 10: Place order
    logger.info("Clicking Place Order button")
    order.click_place_order()

    # Step 11: Verify thank you message
    logger.info("Verifying order success message")
    success_text = order.get_success_message()

    logger.info(f"Success Message Received: {success_text}")

    assert "Thank you. Your order has been received." in success_text, \
        "Order confirmation message not displayed"

    logger.info("========== PLACE ORDER TEST PASSED ==========")