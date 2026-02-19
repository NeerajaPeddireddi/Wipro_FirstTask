# tests/test_login.py
import os
import pytest
from selenium.webdriver.common.by import By

from Capstone_Project.conftest import logger
from Capstone_Project.pages.FindOrders_Byuser import FindOrdersPage
from Capstone_Project.pages.Login import LoginPage

from openpyxl import load_workbook
def read_filterorder_data_excel(file_name):
    base_path = os.path.join(os.path.dirname(__file__), "../data")
    file_path = os.path.join(base_path, file_name)
    logger.info(f"Reading login data from Excel file: {file_path}")
    workbook = load_workbook(file_path)
    sheet = workbook["login_data"]   # Sheet name

    data = []
    headers = []

    # Get headers (first row)
    for cell in sheet[1]:
        headers.append(cell.value)

    # Get rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)
    logger.info(f"Total test data rows loaded: {len(data)}")
    return data


login_test_data = read_filterorder_data_excel("data.xlsx")
@pytest.mark.parametrize("data", login_test_data)
def test_findOrdersByUser(setup,data):
    logger.info("========== STARTING FILTER ORDERS BY USER TEST ==========")
    driver = setup
    find_orders_page = FindOrdersPage(driver)

    # Open My Account page
    logger.info("Clicking on Login/My Account link")
    find_orders_page.click_login()
    current_url = driver.current_url
    logger.info(f"Current URL after clicking login: {current_url}")
    assert "my-account" in driver.current_url.lower(), (
        f"My Account page did not open.\n"
        f"Current URL: {driver.current_url}"
    )
    logger.info(f"Entering Email: {data['email']}")
    find_orders_page.enter_email(data["email"])
    email_value = driver.find_element(By.ID, "username").get_attribute("value")
    logger.info(f"Email entered in field: {email_value}")
    assert email_value == data["email"], (
        f"Email not entered correctly.\nExpected: {data['email']}\nActual: {email_value}"
    )
    logger.info("Entering Password")
    find_orders_page.enter_password(data["password"])
    password_value = driver.find_element(By.ID, "password").get_attribute("value")
    logger.info("Password entered successfully")

    assert password_value == data["password"], "Password not entered correctly"
    logger.info("Selecting Remember Me checkbox")
    find_orders_page.click_checkbox()  # optional
    checkbox_selected = driver.find_element(By.ID, "rememberme").is_selected()
    logger.info(f"Checkbox selected status: {checkbox_selected}")
    assert checkbox_selected, "Checkbox was not selected"
    logger.info("Clicking Login button")
    find_orders_page.click_loginbtn()
    post_login_url = driver.current_url
    logger.info(f"URL after login: {post_login_url}")
    assert "my-account" in driver.current_url.lower(), (
        f"Login failed or wrong page opened.\nCurrent URL: {driver.current_url}"
    )
    logger.info("Checking if Orders table is displayed")
    orders_table=find_orders_page.find_orders()
    assert orders_table.is_displayed(), "Orders table not visible"
    logger.info("Orders table is visible")
    logger.info("========== FILTER ORDERS BY USER TEST PASSED ==========")
