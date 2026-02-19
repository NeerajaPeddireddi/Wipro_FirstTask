import os

import pytest
from selenium.webdriver.common.by import By

from Capstone_Project.conftest import logger
from Capstone_Project.pages.FilterProduct_ByPrice import FilterProduct

from openpyxl import load_workbook

def read_filterproduct_data_excel(file_name):
    base_path = os.path.join(os.path.dirname(__file__), "../data")
    file_path = os.path.join(base_path, file_name)
    logger.info(f"Reading login data from Excel file: {file_path}")
    workbook = load_workbook(file_path)
    sheet = workbook["pricefilter_data"]   # Sheet name

    data = []
    headers = []

    # Get headers from first row
    for cell in sheet[1]:
        headers.append(cell.value)

    # Read remaining rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)
    logger.info(f"Total test data rows loaded: {len(data)}")
    return data


filter_test_data = read_filterproduct_data_excel("data.xlsx")
# --------- Test ----------
@pytest.mark.parametrize("data", filter_test_data)
def test_filter_product(setup, data):
    logger.info("========== STARTING FILTER PRODUCT BY PRICE TEST ==========")
    driver = setup
    filter_page = FilterProduct(driver)
    # Go to Shop
    logger.info("Navigating to Shop page")
    filter_page.go_to_shop()
    current_url = driver.current_url
    logger.info(f"Current URL after navigation: {current_url}")
    assert "shop" in driver.current_url.lower(), (
        f"Shop page did not open.\n"
        f"Current URL: {driver.current_url}"
    )

    min_price = data["min_price"]
    max_price = data["max_price"]
    logger.info(f"Applying price filter -> Min: {min_price}, Max: {max_price}")
    filter_page.filter_by_price(min_price, max_price)
    logger.info("Price filter applied successfully")
    # Check if filtered results appear
    products = driver.find_elements(*filter_page.products)
    product_count = len(products)

    logger.info(f"Number of products displayed after filtering: {product_count}")

    assert len(products) > 0, (
        f"No products found after applying filter: {min_price}-{max_price}"
    )


    # Optional Step 4: Check page title or URL still correct
    updated_url = driver.current_url
    logger.info(f"URL after filtering: {updated_url}")
    assert "shop" in driver.current_url.lower(), (
        f"URL changed unexpectedly after filtering.\nCurrent URL: {driver.current_url}"
    )
    logger.info("========== FILTER PRODUCT TEST PASSED ==========")


