# tests/test_login.py
import os
import pytest

from Capstone_Project.conftest import logger
from Capstone_Project.pages.Login import LoginPage
from openpyxl import load_workbook
def read_login_data_excel(file_name):
    base_path = os.path.join(os.path.dirname(__file__), "../data")
    file_path = os.path.join(base_path, file_name)
    logger.info(f"Reading login data from Excel file: {file_path}")

    workbook = load_workbook(file_path)
    sheet = workbook["login_data"]   # Sheet name

    data = []
    headers = []

    # Get headers (first row)
    for cell in sheet[1]:
        headers.append(cell.value)

    # Get rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)
    logger.info(f"Total test data rows loaded: {len(data)}")
    return data


login_test_data = read_login_data_excel("data.xlsx")
@pytest.mark.order(2)
@pytest.mark.parametrize("data", login_test_data)
def test_login(setup,data):
    logger.info("========== STARTING LOGIN TEST ==========")
    driver = setup
    login_page = LoginPage(driver)
    logger.info("Opening My Account page")
    # Open My Account page
    login_page.click_login()
    assert driver.current_url and "my-account" in driver.current_url.lower(), "Login page did not open"
    logger.info(f"Entering email: {data['email']}")
    login_page.enter_email(data["email"])
    email_value = driver.find_element(*login_page.email_input).get_attribute("value")
    assert email_value == data["email"], f"Email input failed. Expected '{data['email']}', got '{email_value}'"
    logger.info("Entering password")
    login_page.enter_password(data["password"])
    password_value = driver.find_element(*login_page.password_input).get_attribute("value")
    assert password_value == data["password"], "Password input failed"
    logger.info("Clicking checkbox")
    login_page.click_checkbox()  # optional
    checkbox_selected = driver.find_element(*login_page.checkbox).is_selected()
    assert checkbox_selected, "Checkbox was not selected"
    logger.info("Clicking login button")
    login_page.click_loginbtn()
    logger.info("Validating logged in username")
    username_displayed = login_page.get_logged_in_username()
    assert username_displayed == data["username"], f"Expected username '{data['username']}', got '{username_displayed}'"

    # Check login success
    login_page.assert_login_success(data["username"])
    logger.info("========== LOGIN TEST PASSED ==========")
@pytest.mark.order(10)
@pytest.mark.parametrize("data", login_test_data)
def test_logout(setup,data):
    logger.info("========== STARTING LOGOUT TEST ==========")
    driver = setup
    login_page = LoginPage(driver)
    logger.info("Opening My Account page")
    # Open My Account page
    login_page.click_login()
    assert driver.current_url and "my-account" in driver.current_url.lower(), "Login page did not open"
    logger.info("Performing login before logout")
    login_page.enter_email(data["email"])
    email_value = driver.find_element(*login_page.email_input).get_attribute("value")
    assert email_value == data["email"], f"Email input failed. Expected '{data['email']}', got '{email_value}'"

    login_page.enter_password(data["password"])
    password_value = driver.find_element(*login_page.password_input).get_attribute("value")
    assert password_value == data["password"], "Password input failed"
    login_page.click_checkbox()  # optional
    checkbox_selected = driver.find_element(*login_page.checkbox).is_selected()
    assert checkbox_selected, "Checkbox was not selected"
    login_page.click_loginbtn()
    username_displayed = login_page.get_logged_in_username()
    assert username_displayed == data["username"], f"Expected username '{data['username']}', got '{username_displayed}'"

    # Check login success
    login_page.assert_login_success(data["username"])
    logger.info("Clicking sign out")
    # Optionally, sign out after assertion
    login_page.sign_out()
    assert driver.find_element(*login_page.login_btn).is_displayed(), "Logout failed"
    logger.info("========== LOGOUT TEST PASSED ==========")