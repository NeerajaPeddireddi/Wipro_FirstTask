import os

import pytest


from openpyxl.reader.excel import load_workbook

from Capstone_Project.conftest import logger
from Capstone_Project.pages.Registration import RegistrationPage



def read_registration_data_excel(file_name):
    base_path = os.path.join(os.path.dirname(__file__), "../data")
    file_path = os.path.join(base_path, file_name)
    logger.info(f"Reading registration data from Excel file: {file_path}")

    workbook = load_workbook(file_path)
    sheet = workbook["registration_data"]   # Sheet name

    data = []
    headers = []

    # Get headers (first row)
    for cell in sheet[1]:
        headers.append(cell.value)

    # Get rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)
    logger.info(f"Total test data rows loaded: {len(data)}")
    return data
registration_test_data = read_registration_data_excel("data.xlsx")
@pytest.mark.order(1)
@pytest.mark.parametrize("data", registration_test_data)
def test_user_registration(setup,data):
    logger.info("========== STARTING USER REGISTRATION TEST ==========")
    driver = setup
    registration = RegistrationPage(driver)

    # Step 1: Click Sign In
    logger.info("Clicking on Sign Up / My Account link")
    registration.click_sign_up()
    logger.info("Verifying registration page is loaded")
    assert registration.is_registration_page_loaded(), "Registration page did not load"
    # Step 2: Enter Email OR Username
    logger.info(f"Entering email: {data['email']}")
    registration.enter_email(data["email"])
    email_value = driver.find_element(*registration.email_input).get_attribute("value")
    assert email_value == data["email"], f"Email input failed. Expected '{data['email']}', got '{email_value}'"
    logger.info("Entering password")
    registration.enter_password(data["password"])
    password_value = driver.find_element(*registration.password_input).get_attribute("value")
    assert password_value == data["password"], "Password input failed"
    assert driver.find_element(*registration.register).is_enabled(), \
        "Register button is not enabled after entering password"
    # Step 5: Click Register
    logger.info("Clicking Register button")
    registration.click_register()

    driver.implicitly_wait(3)
    # Step 6: Validate Successful Registration
    logger.info("Validating successful registration (Logout link visible)")
    assert registration.is_registration_successful(), \
        "User registration failed"
    logger.info("========== USER REGISTRATION TEST PASSED ==========")

