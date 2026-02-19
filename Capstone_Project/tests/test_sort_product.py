import os
import csv
import pytest

from Capstone_Project.conftest import logger
from Capstone_Project.pages.Sort_Products import SortProducts

from openpyxl import load_workbook
def read_sort_data_excel(file_name):
    base_path = os.path.join(os.path.dirname(__file__), "../data")
    file_path = os.path.join(base_path, file_name)
    logger.info(f"Reading sort data from Excel file: {file_path}")

    workbook = load_workbook(file_path)
    sheet = workbook["productsort_data"]   # Sheet name

    data = []
    headers = []

    # Get headers (first row)
    for cell in sheet[1]:
        headers.append(cell.value)

    # Get rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)
    logger.info(f"Total test data rows loaded: {len(data)}")
    return data


sort_test_data = read_sort_data_excel("data.xlsx")
@pytest.mark.order(7)
@pytest.mark.parametrize("data", sort_test_data)
def test_sort_products(setup, data):
    logger.info("========== STARTING SORT PRODUCTS TEST ==========")
    """
    Test sorting products based on user interest from CSV
    """
    driver = setup
    sort_page = SortProducts(driver)

    # Go to shop page (assuming base URL already handled in setup)
    logger.info("Navigating to Shop page")
    sort_page.go_to_shop()
    current_url = driver.current_url
    logger.info(f"Current URL after navigation: {current_url}")
    assert "shop" in driver.current_url.lower(), (
        f"Shop page did not open.\n"
        f"Current URL: {driver.current_url}"
    )

    # Select sorting option from CSV
    sort_option = data["sort_by"]
    logger.info(f"Selecting sort option: {sort_option}")
    sort_page.select_sort_option(data["sort_by"])


    # Wait a moment for page to reload (optional, adjust if needed)
    driver.implicitly_wait(20)
    current_url = driver.current_url
    logger.info(f"Current URL after sorting: {current_url}")
    # Assert the URL contains the correct 'orderby' parameter

    if data['sort_by'] == "menu_order":
        # Default sorting doesn't change the URL
        logger.info("Validating default sorting (no orderby parameter expected)")
        assert "shop" in current_url
    else:
        # Other sorts append ?orderby=...
        expected_param = f"orderby={data['sort_by']}"
        logger.info(f"Validating URL contains parameter: {expected_param}")
        assert expected_param in current_url, f"Expected '{expected_param}' in URL but got '{current_url}'"
    logger.info("========== SORT PRODUCTS TEST PASSED ==========")
